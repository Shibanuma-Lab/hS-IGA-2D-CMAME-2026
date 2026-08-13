#!/usr/bin/env python3
"""
Collect dynamic sweep post-process outputs into grouped workbooks.

For each velocity and sweep group (rGL/aL/lL/HL), this script reads cases
from results/param_sweep_v{v}_summary.csv and extracts:
1) 4th column of sigmanos_*.csv
2) Last-row valid values (NaN removed) of sigmanos_*.csv
3) K_I_hs / K_I_analytical from step >= 1 in
   J_integral_2D_compare_hs_vs_FEM_*.csv

It also collects the baseline velocity sweep from
results/velocity_baseline_sweep_summary.csv into one workbook by default.

Default output format is XLSX (3 sheets per workbook). If openpyxl is not
available, use --output-format csv as fallback for local validation.
"""

from __future__ import annotations

import argparse
import csv
import math
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np
from core.calnos import analytical_sif


DEFAULT_GROUPS = ("rGL", "aL", "lL", "HL")
ANALYTICAL_HL = 0.05e-3
ANALYTICAL_SIGMA_INF = 1.0e11
ANALYTICAL_EE = 2.06e11
ANALYTICAL_NU = 0.3
ANALYTICAL_RHO = 7800.0


def _parse_csv_list(raw: str) -> List[str]:
    return [x.strip() for x in raw.split(",") if x.strip()]


def _parse_velocities(raw: str) -> List[int]:
    out: List[int] = []
    for tok in _parse_csv_list(raw):
        try:
            out.append(int(tok))
        except ValueError as exc:
            raise ValueError(f"Invalid velocity '{tok}'") from exc
    if not out:
        raise ValueError("No velocities provided.")
    return out


def _parse_groups(raw: str) -> List[str]:
    out = _parse_csv_list(raw)
    if not out:
        raise ValueError("No groups provided.")
    invalid = [g for g in out if g not in DEFAULT_GROUPS]
    if invalid:
        raise ValueError(f"Unsupported groups: {invalid}. Valid: {list(DEFAULT_GROUPS)}")
    return out


def _to_float(value: object) -> Optional[float]:
    if value is None:
        return None
    text = str(value).strip()
    if text == "":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _steps_over_hl(length: int, start_step: int = 1) -> List[float]:
    n = max(0, int(length))
    s0 = int(start_step)
    return [float(s0 + i) for i in range(n)]


def _crack_length_mm_from_count(length: int, start_step: int = 1) -> List[float]:
    steps = _steps_over_hl(length, start_step=start_step)
    return [float(s * ANALYTICAL_HL * 1000.0) for s in steps]


def _read_summary(summary_file: Path, groups: Sequence[str]) -> List[Dict[str, str]]:
    if not summary_file.exists():
        raise FileNotFoundError(f"Summary not found: {summary_file}")

    rows: List[Dict[str, str]] = []
    with open(summary_file, newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            group = (row.get("group") or "").strip()
            if group not in groups:
                continue
            rows.append(row)

    def _idx_key(row: Dict[str, str]) -> int:
        raw = row.get("idx", "")
        try:
            return int(raw)
        except ValueError:
            return 10**9

    rows.sort(key=_idx_key)
    return rows


def _read_velocity_baseline_summary(summary_file: Path) -> List[Dict[str, str]]:
    if not summary_file.exists():
        raise FileNotFoundError(f"Summary not found: {summary_file}")

    rows: List[Dict[str, str]] = []
    with open(summary_file, newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            status = (row.get("status") or "").strip().lower()
            if status not in ("done", "skip_existing"):
                continue
            rows.append(row)

    def _sort_key(row: Dict[str, str]) -> int:
        raw = row.get("v", "")
        try:
            return int(float(raw))
        except ValueError:
            return 10**9

    rows.sort(key=_sort_key)
    return rows


def _find_single_file(folder: Path, pattern: str) -> Optional[Path]:
    matches = sorted(folder.glob(pattern))
    if not matches:
        return None
    if len(matches) > 1:
        print(f"[WARN] Multiple files for pattern '{pattern}' in {folder}; use: {matches[0].name}")
    return matches[0]


def _read_sigmanos(sig_file: Path) -> Tuple[List[float], List[float]]:
    data = np.genfromtxt(sig_file, delimiter=",", dtype=float)

    if data.size == 0:
        return [], []
    if data.ndim == 1:
        data = data.reshape(1, -1)

    # 4th column (index 3)
    col4: List[float] = []
    if data.shape[1] >= 4:
        col4 = [float(v) for v in data[:, 3]]

    # Last row valid (non-NaN) values
    last_row = data[-1, :]
    valid_mask = ~np.isnan(last_row)
    last_valid = [float(v) for v in last_row[valid_mask]]

    return col4, last_valid


def _select_column_key(fieldnames: Sequence[str], target: str) -> Optional[str]:
    stripped_map = {name.strip(): name for name in fieldnames if name is not None}
    if target in stripped_map:
        return stripped_map[target]

    lower_target = target.lower()
    for name in fieldnames:
        if name is None:
            continue
        if name.strip().lower() == lower_target:
            return name
    return None


def _read_ki_norm_hs_over_analytical(j_file: Path, velocity: float) -> Tuple[List[float], List[float]]:
    crack_mm: List[float] = []
    out: List[float] = []
    with open(j_file, newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            return crack_mm, out

        ki_hs_key = _select_column_key(reader.fieldnames, "K_I_hs")
        if ki_hs_key is None:
            raise KeyError(f"Column 'K_I_hs' not found in {j_file}")

        step_key = _select_column_key(reader.fieldnames, "Step")
        if step_key is None:
            raise KeyError(f"Column 'Step' not found in {j_file}")

        for row in reader:
            step = _to_float(row.get(step_key))
            if step is None:
                continue
            step_i = int(round(step))
            if step_i < 1:
                continue

            ki_hs = _to_float(row.get(ki_hs_key))
            if ki_hs is None or math.isnan(ki_hs):
                continue

            ki_analytical = analytical_sif(
                step=step_i,
                V=float(velocity),
                sigma_inf=ANALYTICAL_SIGMA_INF,
                hL=ANALYTICAL_HL,
                EE=ANALYTICAL_EE,
                nu=ANALYTICAL_NU,
                rho=ANALYTICAL_RHO,
            )
            if not np.isfinite(ki_analytical) or abs(float(ki_analytical)) < 1e-14:
                continue

            crack_mm.append(float(step_i * ANALYTICAL_HL * 1000.0))
            out.append(float(ki_hs / ki_analytical))

    return crack_mm, out


def _column_label(row: Dict[str, str]) -> str:
    # folder name is unique and stable.
    return (row.get("folder") or "").strip() or "unknown_case"


def _build_group_payload(
    summary_rows: Sequence[Dict[str, str]],
    results_dir: Path,
    group: str,
) -> Dict[str, Dict[str, List[float]]]:
    sig_col4: Dict[str, List[float]] = {}
    sig_last_valid: Dict[str, List[float]] = {}
    ki_norm: Dict[str, List[float]] = {}

    rows = [r for r in summary_rows if (r.get("group") or "").strip() == group]

    for row in rows:
        folder_name = (row.get("folder") or "").strip()
        if folder_name == "":
            print(f"[WARN] Empty folder field in summary row: {row}")
            continue

        case_dir = results_dir / folder_name
        if not case_dir.exists():
            print(f"[WARN] Case folder missing, skip: {case_dir}")
            continue

        sig_file = _find_single_file(case_dir, "sigmanos_*.csv")
        j_file = _find_single_file(case_dir, "J_integral_2D_compare_hs_vs_FEM_*.csv")

        if sig_file is None:
            print(f"[WARN] sigmanos file missing in {case_dir}")
            continue
        if j_file is None:
            print(f"[WARN] J compare file missing in {case_dir}")
            continue

        label = _column_label(row)
        velocity = _to_float(row.get("v"))
        if velocity is None:
            print(f"[WARN] Missing velocity in summary row, skip: {row}")
            continue

        try:
            col4, last_valid = _read_sigmanos(sig_file)
            ki_crack_mm, ki_vals = _read_ki_norm_hs_over_analytical(j_file, velocity=float(velocity))
        except Exception as exc:
            print(f"[WARN] Failed to parse {case_dir.name}: {type(exc).__name__}: {exc}")
            continue

        sig_col4[f"{label} | crack_length_mm"] = _crack_length_mm_from_count(len(col4), start_step=1)
        sig_col4[f"{label} | sigmanos_col4"] = col4
        sig_last_valid[label] = last_valid
        ki_norm[f"{label} | crack_length_mm"] = ki_crack_mm
        ki_norm[f"{label} | K_I_hS/K_I_analytical"] = ki_vals

    return {
        "sigmanos_col4": sig_col4,
        "sigmanos_last_valid": sig_last_valid,
        "ki_norm_step1_to_end": ki_norm,
    }


def _build_velocity_baseline_payload(
    summary_rows: Sequence[Dict[str, str]],
    results_dir: Path,
) -> Dict[str, Dict[str, List[float]]]:
    sig_col4: Dict[str, List[float]] = {}
    sig_last_valid: Dict[str, List[float]] = {}
    ki_norm: Dict[str, List[float]] = {}

    for row in summary_rows:
        folder_name = (row.get("folder") or "").strip()
        if folder_name == "":
            print(f"[WARN] Empty folder field in velocity baseline row: {row}")
            continue

        case_dir = results_dir / folder_name
        if not case_dir.exists():
            print(f"[WARN] Case folder missing, skip: {case_dir}")
            continue

        sig_file = _find_single_file(case_dir, "sigmanos_*.csv")
        j_file = _find_single_file(case_dir, "J_integral_2D_compare_hs_vs_FEM_*.csv")

        if sig_file is None:
            print(f"[WARN] sigmanos file missing in {case_dir}")
            continue
        if j_file is None:
            print(f"[WARN] J compare file missing in {case_dir}")
            continue

        label = f"v={int(float(row.get('v', 0)))}"
        velocity = _to_float(row.get("v"))
        if velocity is None:
            print(f"[WARN] Missing velocity in velocity baseline row, skip: {row}")
            continue

        try:
            col4, last_valid = _read_sigmanos(sig_file)
            ki_crack_mm, ki_vals = _read_ki_norm_hs_over_analytical(j_file, velocity=float(velocity))
        except Exception as exc:
            print(f"[WARN] Failed to parse {case_dir.name}: {type(exc).__name__}: {exc}")
            continue

        sig_col4[f"{label} | crack_length_mm"] = _crack_length_mm_from_count(len(col4), start_step=1)
        sig_col4[f"{label} | sigmanos_col4"] = col4
        sig_last_valid[label] = last_valid
        ki_norm[f"{label} | crack_length_mm"] = ki_crack_mm
        ki_norm[f"{label} | K_I_hS/K_I_analytical"] = ki_vals

    return {
        "sigmanos_col4": sig_col4,
        "sigmanos_last_valid": sig_last_valid,
        "ki_norm_step1_to_end": ki_norm,
    }


def _write_columns_to_csv(csv_file: Path, columns: Dict[str, List[float]]) -> None:
    csv_file.parent.mkdir(parents=True, exist_ok=True)
    headers = list(columns.keys())
    max_len = max((len(v) for v in columns.values()), default=0)

    with open(csv_file, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for i in range(max_len):
            row = []
            for h in headers:
                vals = columns[h]
                row.append(vals[i] if i < len(vals) else "")
            writer.writerow(row)


def _write_csv_bundle(base_path: Path, payload: Dict[str, Dict[str, List[float]]]) -> None:
    out_dir = base_path.with_suffix("")
    out_dir.mkdir(parents=True, exist_ok=True)
    for sheet_name, columns in payload.items():
        csv_file = out_dir / f"{sheet_name}.csv"
        _write_columns_to_csv(csv_file, columns)


def _write_xlsx(workbook_file: Path, payload: Dict[str, Dict[str, List[float]]]) -> None:
    try:
        from openpyxl import Workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "openpyxl is required for XLSX output. Install it with: pip install openpyxl"
        ) from exc

    workbook_file.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    first = True

    for sheet_name, columns in payload.items():
        title = sheet_name[:31]
        if first:
            ws = wb.active
            ws.title = title
            first = False
        else:
            ws = wb.create_sheet(title=title)

        headers = list(columns.keys())
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
            vals = columns[header]
            for row_idx, val in enumerate(vals, start=2):
                if isinstance(val, float) and math.isnan(val):
                    ws.cell(row=row_idx, column=col_idx, value=None)
                else:
                    ws.cell(row=row_idx, column=col_idx, value=float(val))

    wb.save(workbook_file)


def main() -> int:
    parser = argparse.ArgumentParser(description="Collect dynamic sweep outputs into grouped workbooks.")
    parser.add_argument("--results-dir", type=Path, default=Path("results"), help="Results root directory.")
    parser.add_argument("--output-dir", type=Path, default=Path("results") / "organized", help="Output directory.")
    parser.add_argument("--velocities", type=str, default="500,1000", help="Comma-separated velocities.")
    parser.add_argument("--groups", type=str, default=",".join(DEFAULT_GROUPS), help="Comma-separated groups.")
    parser.add_argument(
        "--skip-velocity-baseline",
        action="store_true",
        help="Do not collect results/velocity_baseline_sweep_summary.csv.",
    )
    parser.add_argument(
        "--output-format",
        choices=("xlsx", "csv"),
        default="xlsx",
        help="xlsx: one workbook per group; csv: one folder of 3 csv files per group.",
    )
    args = parser.parse_args()

    results_dir = args.results_dir.resolve()
    output_dir = args.output_dir.resolve()
    velocities = _parse_velocities(args.velocities)
    groups = _parse_groups(args.groups)

    if not results_dir.exists():
        raise FileNotFoundError(f"Results directory not found: {results_dir}")

    print(f"[INFO] results_dir={results_dir}")
    print(f"[INFO] output_dir={output_dir}")
    print(f"[INFO] velocities={velocities}")
    print(f"[INFO] groups={groups}")
    print(f"[INFO] output_format={args.output_format}")

    for v in velocities:
        summary_file = results_dir / f"param_sweep_v{v}_summary.csv"
        try:
            summary_rows = _read_summary(summary_file, groups)
        except FileNotFoundError as exc:
            print(f"[WARN] {exc}")
            continue

        if not summary_rows:
            print(f"[WARN] No rows found in summary for v={v}: {summary_file}")
            continue

        for group in groups:
            payload = _build_group_payload(summary_rows, results_dir, group)
            ncols = len(payload["sigmanos_col4"])
            if ncols == 0:
                print(f"[WARN] No valid cases for v={v}, group={group}; skip output.")
                continue

            base_name = f"v{v}_fix_{group}"
            if args.output_format == "xlsx":
                out_file = output_dir / f"{base_name}.xlsx"
                _write_xlsx(out_file, payload)
                print(f"[OK] {out_file} (columns={ncols})")
            else:
                out_file = output_dir / f"{base_name}.csv"
                _write_csv_bundle(out_file, payload)
                print(f"[OK] {out_file.with_suffix('')}/*.csv (columns={ncols})")

    if not args.skip_velocity_baseline:
        baseline_summary = results_dir / "velocity_baseline_sweep_summary.csv"
        try:
            baseline_rows = _read_velocity_baseline_summary(baseline_summary)
        except FileNotFoundError as exc:
            print(f"[WARN] {exc}")
            baseline_rows = []

        if baseline_rows:
            payload = _build_velocity_baseline_payload(baseline_rows, results_dir)
            ncols = len(payload["sigmanos_col4"])
            if ncols == 0:
                print("[WARN] No valid velocity-baseline cases; skip output.")
            elif args.output_format == "xlsx":
                out_file = output_dir / "velocity_baseline.xlsx"
                _write_xlsx(out_file, payload)
                print(f"[OK] {out_file} (columns={ncols})")
            else:
                out_file = output_dir / "velocity_baseline.csv"
                _write_csv_bundle(out_file, payload)
                print(f"[OK] {out_file.with_suffix('')}/*.csv (columns={ncols})")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
