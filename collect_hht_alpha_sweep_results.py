#!/usr/bin/env python3
"""
Collect HHT-alpha sweep outputs into velocity-grouped workbooks.

The script reads rows from results/hht_alpha_sweep_summary.csv and extracts:
1) 4th column of sigmanos_*.csv
2) Last-row valid values (NaN removed) of sigmanos_*.csv
3) K_I_norm_hs_over_fem column from step >= 1 in
   J_integral_2D_compare_hs_vs_FEM_*.csv
"""

from __future__ import annotations

import argparse
import csv
import math
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np


def _parse_csv_list(raw: str) -> List[str]:
    return [x.strip() for x in str(raw).split(",") if x.strip()]


def _parse_velocities(raw: str) -> List[int]:
    out: List[int] = []
    for tok in _parse_csv_list(raw):
        try:
            out.append(int(tok))
        except ValueError as exc:
            raise ValueError(f"Invalid velocity '{tok}'") from exc
    return list(dict.fromkeys(out))


def _parse_statuses(raw: str) -> List[str]:
    vals = [x.lower() for x in _parse_csv_list(raw)]
    return list(dict.fromkeys(vals))


def _to_float(value: object) -> Optional[float]:
    if value is None:
        return None
    text = str(value).strip()
    if text == "":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _to_int(value: object) -> Optional[int]:
    val = _to_float(value)
    if val is None:
        return None
    return int(val)


def _read_summary(
    summary_file: Path,
    velocities: Sequence[int],
    statuses: Sequence[str],
) -> List[Dict[str, str]]:
    if not summary_file.exists():
        raise FileNotFoundError(f"Summary not found: {summary_file}")

    velocities_set = set(int(v) for v in velocities)
    statuses_set = set(s.lower() for s in statuses)

    rows: List[Dict[str, str]] = []
    with open(summary_file, newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            v = _to_int(row.get("v"))
            if v is None:
                continue
            if velocities_set and v not in velocities_set:
                continue

            status = str(row.get("status") or "").strip().lower()
            if statuses_set and status not in statuses_set:
                continue

            rows.append(row)

    def _sort_key(row: Dict[str, str]) -> Tuple[int, float, int]:
        v = _to_int(row.get("v"))
        alpha = _to_float(row.get("alpha"))
        idx = _to_int(row.get("idx"))
        return (
            int(v) if v is not None else 10**9,
            float(alpha) if alpha is not None else float("inf"),
            int(idx) if idx is not None else 10**9,
        )

    rows.sort(key=_sort_key)
    return rows


def _find_single_file(case_dir: Path, pattern: str) -> Optional[Path]:
    matches = sorted(case_dir.glob(pattern))
    if not matches:
        matches = sorted(case_dir.glob(f"excel/{pattern}"))
    if not matches:
        matches = sorted(case_dir.rglob(pattern))
    if not matches:
        return None
    if len(matches) > 1:
        print(f"[WARN] Multiple files for pattern '{pattern}' in {case_dir}; use: {matches[0]}")
    return matches[0]


def _read_sigmanos(sig_file: Path) -> Tuple[List[float], List[float]]:
    data = np.genfromtxt(sig_file, delimiter=",", dtype=float)
    if data.size == 0:
        return [], []
    if data.ndim == 1:
        data = data.reshape(1, -1)

    col4: List[float] = []
    if data.shape[1] >= 4:
        col4 = [float(v) for v in data[:, 3]]

    last_row = data[-1, :]
    valid_mask = ~np.isnan(last_row)
    last_valid = [float(v) for v in last_row[valid_mask]]
    return col4, last_valid


def _select_column_key(fieldnames: Sequence[str], target: str) -> Optional[str]:
    stripped_map = {name.strip(): name for name in fieldnames if name is not None}
    if target in stripped_map:
        return stripped_map[target]

    lower_target = target.lower()
    for name in fieldnames:
        if name is None:
            continue
        if name.strip().lower() == lower_target:
            return name
    return None


def _read_ki_norm(j_file: Path) -> List[float]:
    out: List[float] = []
    with open(j_file, newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            return out

        ki_key = _select_column_key(reader.fieldnames, "K_I_norm_hs_over_fem")
        if ki_key is None:
            raise KeyError(f"Column 'K_I_norm_hs_over_fem' not found in {j_file}")

        step_key = _select_column_key(reader.fieldnames, "Step")
        for row in reader:
            ki = _to_float(row.get(ki_key))
            if ki is None or math.isnan(ki):
                continue

            if step_key is not None:
                step = _to_float(row.get(step_key))
                if step is None or step < 1.0:
                    continue

            out.append(float(ki))
    return out


def _column_label(row: Dict[str, str]) -> str:
    alpha = _to_float(row.get("alpha"))
    if alpha is None:
        return (row.get("folder") or "").strip() or "unknown_case"
    return f"alpha={alpha:+.3f}"


def _make_unique_label(base: str, used: Dict[str, int]) -> str:
    if base not in used:
        used[base] = 1
        return base
    used[base] += 1
    return f"{base}#{used[base]}"


def _build_velocity_payload(
    velocity_rows: Sequence[Dict[str, str]],
    results_dir: Path,
) -> Dict[str, Dict[str, List[float]]]:
    sig_col4: Dict[str, List[float]] = {}
    sig_last_valid: Dict[str, List[float]] = {}
    ki_norm: Dict[str, List[float]] = {}

    used_labels: Dict[str, int] = {}
    for row in velocity_rows:
        folder_name = (row.get("folder") or "").strip()
        if folder_name == "":
            print(f"[WARN] Empty folder field in summary row: {row}")
            continue

        case_dir = results_dir / folder_name
        if not case_dir.exists():
            print(f"[WARN] Case folder missing, skip: {case_dir}")
            continue

        sig_file = _find_single_file(case_dir, "sigmanos_*.csv")
        j_file = _find_single_file(case_dir, "J_integral_2D_compare_hs_vs_FEM_*.csv")
        if sig_file is None:
            print(f"[WARN] sigmanos file missing in {case_dir}")
            continue
        if j_file is None:
            print(f"[WARN] J compare file missing in {case_dir}")
            continue

        label = _make_unique_label(_column_label(row), used_labels)

        try:
            col4, last_valid = _read_sigmanos(sig_file)
            ki_vals = _read_ki_norm(j_file)
        except Exception as exc:
            print(f"[WARN] Failed to parse {case_dir.name}: {type(exc).__name__}: {exc}")
            continue

        sig_col4[label] = col4
        sig_last_valid[label] = last_valid
        ki_norm[label] = ki_vals

    return {
        "sigmanos_col4": sig_col4,
        "sigmanos_last_valid": sig_last_valid,
        "ki_norm_step1_to_end": ki_norm,
    }


def _write_columns_to_csv(csv_file: Path, columns: Dict[str, List[float]]) -> None:
    csv_file.parent.mkdir(parents=True, exist_ok=True)
    headers = list(columns.keys())
    max_len = max((len(v) for v in columns.values()), default=0)

    with open(csv_file, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for i in range(max_len):
            row = []
            for h in headers:
                vals = columns[h]
                row.append(vals[i] if i < len(vals) else "")
            writer.writerow(row)


def _write_csv_bundle(base_path: Path, payload: Dict[str, Dict[str, List[float]]]) -> None:
    out_dir = base_path.with_suffix("")
    out_dir.mkdir(parents=True, exist_ok=True)
    for sheet_name, columns in payload.items():
        csv_file = out_dir / f"{sheet_name}.csv"
        _write_columns_to_csv(csv_file, columns)


def _write_xlsx(workbook_file: Path, payload: Dict[str, Dict[str, List[float]]]) -> None:
    try:
        from openpyxl import Workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "openpyxl is required for XLSX output. Install it with: pip install openpyxl"
        ) from exc

    workbook_file.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    first = True

    for sheet_name, columns in payload.items():
        title = sheet_name[:31]
        if first:
            ws = wb.active
            ws.title = title
            first = False
        else:
            ws = wb.create_sheet(title=title)

        headers = list(columns.keys())
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
            vals = columns[header]
            for row_idx, val in enumerate(vals, start=2):
                if isinstance(val, float) and math.isnan(val):
                    ws.cell(row=row_idx, column=col_idx, value=None)
                else:
                    ws.cell(row=row_idx, column=col_idx, value=float(val))

    wb.save(workbook_file)


def _collect_velocity_ids(rows: Iterable[Dict[str, str]]) -> List[int]:
    out: List[int] = []
    seen = set()
    for row in rows:
        v = _to_int(row.get("v"))
        if v is None or v in seen:
            continue
        seen.add(v)
        out.append(int(v))
    return sorted(out)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Collect HHT-alpha sweep outputs into velocity-grouped workbooks."
    )
    parser.add_argument("--results-dir", type=Path, default=Path("results"), help="Results root directory.")
    parser.add_argument(
        "--summary-file",
        type=Path,
        default=Path("results") / "hht_alpha_sweep_summary.csv",
        help="Input summary CSV generated by run_hht_alpha_sweep.py.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("results") / "organized_hht_alpha",
        help="Output directory.",
    )
    parser.add_argument(
        "--velocities",
        type=str,
        default="",
        help="Comma-separated velocities. Empty means use all velocities in summary.",
    )
    parser.add_argument(
        "--statuses",
        type=str,
        default="done,skip_existing",
        help="Comma-separated statuses to include; empty means include all.",
    )
    parser.add_argument(
        "--output-format",
        choices=("xlsx", "csv"),
        default="xlsx",
        help="xlsx: one workbook per velocity; csv: one folder of 3 csv files per velocity.",
    )
    args = parser.parse_args()

    results_dir = args.results_dir.resolve()
    summary_file = args.summary_file.resolve()
    output_dir = args.output_dir.resolve()
    velocities = _parse_velocities(args.velocities)
    statuses = _parse_statuses(args.statuses)

    if not results_dir.exists():
        raise FileNotFoundError(f"Results directory not found: {results_dir}")

    print(f"[INFO] results_dir={results_dir}")
    print(f"[INFO] summary_file={summary_file}")
    print(f"[INFO] output_dir={output_dir}")
    print(f"[INFO] velocities={velocities if velocities else 'ALL'}")
    print(f"[INFO] statuses={statuses if statuses else 'ALL'}")
    print(f"[INFO] output_format={args.output_format}")

    rows = _read_summary(summary_file, velocities=velocities, statuses=statuses)
    if not rows:
        print("[WARN] No summary rows matched filters. Nothing to do.")
        return 0

    velocity_ids = _collect_velocity_ids(rows)
    for v in velocity_ids:
        v_rows = [r for r in rows if _to_int(r.get("v")) == v]
        payload = _build_velocity_payload(v_rows, results_dir)
        ncols = len(payload["sigmanos_col4"])
        if ncols == 0:
            print(f"[WARN] No valid cases for v={v}; skip output.")
            continue

        base_name = f"hht_alpha_v{v}"
        if args.output_format == "xlsx":
            out_file = output_dir / f"{base_name}.xlsx"
            _write_xlsx(out_file, payload)
            print(f"[OK] {out_file} (columns={ncols})")
        else:
            out_file = output_dir / f"{base_name}.csv"
            _write_csv_bundle(out_file, payload)
            print(f"[OK] {out_file.with_suffix('')}/*.csv (columns={ncols})")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
